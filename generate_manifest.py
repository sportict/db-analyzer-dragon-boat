#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
manifest.json を data/ 配下のCSV/XLSXから自動生成するスクリプト。

特徴:
  * data/<大会日付>/ 配下に置かれたCSV/XLSX を一覧化
  * XLSX があれば自動的にCSV(SPI ProX2 形式互換)に変換
  * 速度の単位を m/s に正規化(XLSX の km/h を変換)
  * ファイル名規約: R<番号>_<クラス>_<レーン>_<チーム名>.csv|xlsx

依存:
  pip install openpyxl

使い方:
  python3 generate_manifest.py
"""
import json
import os
import re
import sys
from datetime import date, datetime, time, timedelta

try:
    import openpyxl
except ImportError:
    openpyxl = None

ROOT = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(ROOT, "data")

EVENT_LABELS = {
    "2025-10-05": "2025/10/05 第19回 KIX国際交流ドラゴンボート大会 (250m)",
    "2025-11-02": "2025/11/02 天神祭奉納 日本国際選手権 (250m)",
    "2025-11-16": "2025/11/16 第13回 スモールドラゴンボート日本選手権 (250m)",
}

FNAME_RE = re.compile(r"^(R\d+)_([^_]+)_(\d+)_(.+?)\.(csv|xlsx)$")


def parse_filename(fn: str):
    m = FNAME_RE.match(fn)
    if not m:
        return None
    rno, klass, _lane, team, _ext = m.groups()
    return {"race": f"{rno} {klass}", "team": team}


# ========== XLSX → CSV 変換 ==========
def _fmt_cumulative(v):
    """time / timedelta / float(Excel日付) を 'MM:SS.fff' 形式に"""
    if v is None:
        return ""
    if isinstance(v, time):
        total = v.hour * 3600 + v.minute * 60 + v.second + v.microsecond / 1e6
    elif isinstance(v, timedelta):
        total = v.total_seconds()
    elif isinstance(v, datetime):
        total = v.hour * 3600 + v.minute * 60 + v.second + v.microsecond / 1e6
    else:
        try:
            total = float(v) * 86400.0  # Excel fraction of day
        except Exception:
            return str(v)
    mm = int(total // 60)
    ss = total - mm * 60
    return f"{mm:02d}:{ss:06.3f}"


def _fmt_time(v):
    if isinstance(v, datetime):
        return v.strftime("%M:%S.") + f"{v.microsecond // 1000:03d}"
    return str(v) if v is not None else ""


def convert_xlsx_to_csv(xlsx_path: str, csv_path: str) -> bool:
    """xlsx を SPI ProX2 形式の CSV に変換。速度を m/s に正規化。"""
    if openpyxl is None:
        print(f"[skip] openpyxl 未インストール: {xlsx_path}", file=sys.stderr)
        return False

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # ヘッダ行 (Time, Cumulative Time, ...) を検索
    header_row_idx = None
    header = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row and row[0] == "Time":
            header_row_idx = i
            header = list(row)
            break
    if header_row_idx is None:
        print(f"[error] ヘッダ行が見つかりません: {xlsx_path}", file=sys.stderr)
        return False

    # 速度列のインデックスとスケール
    speed_col = None
    speed_factor = 1.0  # m/s
    for idx, name in enumerate(header):
        if name and isinstance(name, str) and name.startswith("Speed"):
            speed_col = idx
            if "km/h" in name:
                speed_factor = 1.0 / 3.6  # km/h → m/s
            header[idx] = "Speed (m/s)"
            break

    # Filtered Velocity 列も対応(同様)
    for idx, name in enumerate(header):
        if name and isinstance(name, str) and name.startswith("Filtered Velocity"):
            if "km/h" in name:
                header[idx] = "Filtered Velocity (m/s)"
            break

    # CSV 出力
    with open(csv_path, "w", encoding="utf-8") as f:
        # 上部メタ行 (ヘッダー行までを保持)
        for i in range(1, header_row_idx):
            row = next(ws.iter_rows(min_row=i, max_row=i, values_only=True))
            cells = ["" if c is None else (str(c) if not isinstance(c, datetime) else c.isoformat()) for c in row]
            f.write(",".join(cells) + "\n")

        # ヘッダ行
        f.write(",".join("" if h is None else str(h) for h in header) + "\n")

        # データ行
        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if all(c is None or c == "" for c in row):
                continue
            cells = []
            for ci, c in enumerate(row):
                if c is None:
                    cells.append("")
                elif ci == 0:  # Time
                    cells.append(_fmt_time(c) if isinstance(c, datetime) else str(c))
                elif ci == 1:  # Cumulative Time
                    cells.append(_fmt_cumulative(c))
                elif ci == speed_col:
                    try:
                        cells.append(f"{float(c) * speed_factor:.3f}")
                    except Exception:
                        cells.append(str(c))
                elif isinstance(c, float):
                    cells.append(f"{c:.4f}".rstrip("0").rstrip(".") or "0")
                else:
                    cells.append(str(c))
            f.write(",".join(cells) + "\n")

    return True


# ========== マニフェスト構築 ==========
def build_event(event_dir: str, event_id: str):
    name = EVENT_LABELS.get(event_id, event_id)
    races = []
    converted = 0
    files = sorted(os.listdir(event_dir))

    # XLSX を CSV に変換
    for fn in files:
        if fn.endswith(".xlsx") and not fn.startswith("~$"):
            csv_name = fn[:-5] + ".csv"
            csv_path = os.path.join(event_dir, csv_name)
            xlsx_path = os.path.join(event_dir, fn)
            if not os.path.exists(csv_path) or os.path.getmtime(xlsx_path) > os.path.getmtime(csv_path):
                if convert_xlsx_to_csv(xlsx_path, csv_path):
                    converted += 1

    # 改めて一覧化(CSVのみ)
    files = sorted(os.listdir(event_dir))
    for fn in files:
        if not fn.endswith(".csv"):
            continue
        meta = parse_filename(fn) or {"race": fn, "team": ""}
        races.append({
            "file": f"data/{event_id}/{fn}",
            "team": meta["team"],
            "race": meta["race"],
        })

    if converted:
        print(f"  [convert] {event_id}: xlsx→csv {converted}件")

    return {"name": name, "id": event_id, "races": races}


def main():
    if not os.path.isdir(DATA_DIR):
        print(f"[error] data/ ディレクトリが見つかりません: {DATA_DIR}", file=sys.stderr)
        sys.exit(1)

    events = []
    for entry in sorted(os.listdir(DATA_DIR)):
        d = os.path.join(DATA_DIR, entry)
        if not os.path.isdir(d):
            continue
        ev = build_event(d, entry)
        if ev["races"]:
            events.append(ev)

    manifest = {
        "title": "ドラゴンボート レース GPS データ",
        "updated": date.today().isoformat(),
        "events": events,
    }

    out = os.path.join(ROOT, "manifest.json")
    with open(out, "w", encoding="utf-8") as f:
        json.dump(manifest, f, ensure_ascii=False, indent=2)

    total = sum(len(e["races"]) for e in events)
    print(f"[ok] {out} を生成しました ({len(events)} 大会 / {total} レース)")


if __name__ == "__main__":
    main()
